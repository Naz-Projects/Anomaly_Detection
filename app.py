"""
Anomaly Detector - Manufacturing Test Results Analysis
A Streamlit application for detecting anomalies in manufacturing test data.
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from src.data_loader import DataLoader
from src.detector import AnomalyDetector

# Page configuration
st.set_page_config(
    page_title="Anomaly Detector",
    page_icon=":mag:",
    layout="wide"
)

# Initialize session state
if 'df' not in st.session_state:
    st.session_state.df = None
if 'selected_item' not in st.session_state:
    st.session_state.selected_item = None
if 'filters' not in st.session_state:
    st.session_state.filters = []
if 'analysis_results' not in st.session_state:
    st.session_state.analysis_results = None


def add_filter():
    """Add a new empty filter to the session state."""
    st.session_state.filters.append({'result_name': None, 'lower_bound': 0.0, 'upper_bound': 0.0})


def remove_filter(index):
    """Remove a filter at the specified index."""
    st.session_state.filters.pop(index)


def create_formatted_excel(df: pd.DataFrame) -> BytesIO:
    """
    Create an Excel file with conditional formatting for abnormal rows.

    Parameters:
    -----------
    df : pd.DataFrame
        The results dataframe with IS_OUTLIER column

    Returns:
    --------
    BytesIO object containing the formatted Excel file
    """
    # Create a BytesIO buffer
    output = BytesIO()

    # Write dataframe to Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')

    # Load the workbook to apply formatting
    output.seek(0)
    workbook = load_workbook(output)
    worksheet = workbook['Results']

    # Define light red fill for abnormal rows
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

    # Find the IS_OUTLIER column index
    header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    is_outlier_col_idx = header_row.index('IS_OUTLIER') + 1 if 'IS_OUTLIER' in header_row else None

    if is_outlier_col_idx:
        # Iterate through rows and apply formatting
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            # Check if IS_OUTLIER is ABNORMAL
            if row[is_outlier_col_idx - 1].value == 'ABNORMAL':
                # Apply red fill to all cells in the row
                for cell in row:
                    cell.fill = red_fill

    # Save to buffer
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def display_results(result_df: pd.DataFrame):
    """Display analysis results with metrics and tables."""

    st.divider()
    st.subheader("Analysis Results")

    # Get summary statistics
    stats = AnomalyDetector.get_summary_stats(result_df)

    # Display summary metrics
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Analyzed", stats['total_analyzed'])
    with col2:
        st.metric("Normal", stats['normal_count'])
    with col3:
        st.metric("Abnormal", stats['abnormal_count'])

    # Check if there are anomalies
    abnormal_df = result_df[result_df['IS_OUTLIER'] == 'ABNORMAL'].copy()

    if not abnormal_df.empty:
        # Detailed results table - only show rows with anomalies
        st.markdown("### Detailed Anomaly Records")

        # Select relevant columns for display
        display_cols = ['ITEM_NUMBER', 'TEST_NUMBER', 'RESULT_NAME', 'RESPONSE', 'Lower_Bound', 'Upper_Bound', 'IS_OUTLIER']
        display_df = abnormal_df[display_cols]

        st.dataframe(display_df, use_container_width=True, hide_index=True)

        # Download button for full results
        st.markdown("### Export Results")

        # Create formatted Excel file
        excel_file = create_formatted_excel(result_df)

        st.download_button(
            label="Download Results (Excel with highlighting)",
            data=excel_file,
            file_name="anomaly_detection_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False
        )
    else:
        st.success("No anomalies detected! All measurements are within acceptable ranges.")

    # Clear results button
    if st.button("Clear Results"):
        st.session_state.analysis_results = None
        st.rerun()


def main():
    """Main application entry point."""

    st.title("Anomaly Detector")
    st.caption("Automated detection of abnormal test results")
    st.divider()

    # File upload section
    st.subheader("Upload Test Results")

    col_upload, col_stats = st.columns([1, 2])

    with col_upload:
        # File uploader
        uploaded_file = st.file_uploader(
            "Choose an Excel file",
            type=['xlsx', 'xls'],
            help="Upload your manufacturing test results in Excel format"
        )

    if uploaded_file is not None:
        # Load the file
        with st.spinner("Loading file..."):
            df, error = DataLoader.load_excel(uploaded_file)

        if error:
            st.error(f":x: {error}")
            return

        # Store in session state
        st.session_state.df = df

        # Display basic statistics in the same row
        with col_stats:
            st.success(f"File loaded: {uploaded_file.name}")
            stats = DataLoader.get_basic_stats(df)
            metric_col1, metric_col2, metric_col3 = st.columns(3)
            with metric_col1:
                st.metric("Total Rows", f"{stats['total_rows']:,}")
            with metric_col2:
                st.metric("Products", stats['total_items'])
            with metric_col3:
                st.metric("Tests", stats['total_tests'])

        # Data preview
        with st.expander("Preview Data", expanded=False):
            st.dataframe(df.head(10), use_container_width=True)

        st.divider()

        # Product selection - vertical but compact with multi-select
        st.markdown("**Select Products**")
        col1, col2 = st.columns([2, 3])
        with col1:
            item_numbers = DataLoader.get_item_numbers(df)

            # Multi-select with all products selected by default
            selected_items = st.multiselect(
                "Choose ITEM_NUMBER(s)",
                options=item_numbers,
                default=item_numbers,  # All products selected by default
                help="Select one or more products to analyze",
                label_visibility="collapsed"
            )

        if selected_items:
            st.session_state.selected_item = selected_items

            # Show test count for selected items
            total_tests = sum([DataLoader.get_test_count(df, item) for item in selected_items])
            with col2:
                st.markdown(f"<br>**{len(selected_items)}** product(s), **{total_tests}** test sessions", unsafe_allow_html=True)

            st.divider()

            # Configure detection criteria
            st.subheader("Detection Criteria")
            st.caption("Add filters to detect anomalies. Values outside the ranges will be flagged.")

            # Get analyzable result names across all selected products
            result_names = []
            for item in selected_items:
                item_results = DataLoader.get_analyzable_result_names(df, item)
                result_names.extend(item_results)
            result_names = sorted(list(set(result_names)))  # Remove duplicates and sort

            if not result_names:
                st.warning(":warning: No analyzable test types found for selected products")
                return

            st.caption(f"{len(result_names)} test types available for analysis")

            # Display existing filters
            if st.session_state.filters:
                for i, filter_item in enumerate(st.session_state.filters):
                    col1, col2, col3, col4 = st.columns([3, 1.5, 1.5, 0.5])

                    with col1:
                        selected_result = st.selectbox(
                            "Test Type",
                            options=result_names,
                            key=f"result_name_{i}",
                            index=result_names.index(filter_item['result_name']) if filter_item['result_name'] in result_names else 0
                        )
                        st.session_state.filters[i]['result_name'] = selected_result

                        # Show current data range across all selected products
                        all_min_vals = []
                        all_max_vals = []
                        for item in selected_items:
                            min_val, max_val = DataLoader.get_value_range(df, item, selected_result)
                            if min_val is not None:
                                all_min_vals.append(min_val)
                            if max_val is not None:
                                all_max_vals.append(max_val)

                        if all_min_vals and all_max_vals:
                            overall_min = min(all_min_vals)
                            overall_max = max(all_max_vals)
                            st.caption(f"Range: {overall_min:.3f} to {overall_max:.3f}")

                    with col2:
                        lower = st.number_input(
                            "Lower",
                            value=filter_item['lower_bound'],
                            format="%.3f",
                            key=f"lower_{i}"
                        )
                        st.session_state.filters[i]['lower_bound'] = lower

                    with col3:
                        upper = st.number_input(
                            "Upper",
                            value=filter_item['upper_bound'],
                            format="%.3f",
                            key=f"upper_{i}"
                        )
                        st.session_state.filters[i]['upper_bound'] = upper

                    with col4:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("✕", key=f"remove_{i}", help="Remove"):
                            remove_filter(i)
                            st.rerun()

            # Buttons in a compact row
            st.markdown("")
            btn_col1, btn_col2 = st.columns([1, 3])
            with btn_col1:
                if st.button("➕ Add Filter", use_container_width=True):
                    add_filter()
                    st.rerun()
            with btn_col2:
                if st.button("▶ Run Analysis", type="primary", use_container_width=True):
                    if st.session_state.filters:
                        # Build criteria dictionary from filters
                        criteria = {
                            f['result_name']: (f['lower_bound'], f['upper_bound'])
                            for f in st.session_state.filters
                            if f['result_name'] is not None
                        }

                        # Run anomaly detection for all selected products
                        with st.spinner("Analyzing data..."):
                            all_results = []
                            for item in selected_items:
                                result_df = AnomalyDetector.detect_anomalies(
                                    st.session_state.df,
                                    item,
                                    criteria
                                )
                                all_results.append(result_df)

                            # Combine all results
                            combined_results = pd.concat(all_results, ignore_index=True)
                            st.session_state.analysis_results = combined_results

                        st.success(f"Analysis complete! Processed {len(selected_items)} product(s) with {len(st.session_state.filters)} filter(s)")
                        st.rerun()
                    else:
                        st.warning("Please add at least one filter to run the analysis")

            # Display results if available
            if st.session_state.analysis_results is not None:
                display_results(st.session_state.analysis_results)

    else:
        st.info(":point_up: Please upload an Excel file to get started")


if __name__ == "__main__":
    main()
