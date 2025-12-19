"""Excel export functionality with conditional formatting."""

import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def create_formatted_excel(df: pd.DataFrame) -> BytesIO:
    """
    Create an Excel file with conditional formatting for abnormal rows.

    Parameters:
    -----------
    df : pd.DataFrame
        The results dataframe with IS_OUTLIER column

    Returns:
    --------
    BytesIO object containing the formatted Excel file
    """
    # Create a BytesIO buffer
    output = BytesIO()

    # Write dataframe to Excel
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')

    # Load the workbook to apply formatting
    output.seek(0)
    workbook = load_workbook(output)
    worksheet = workbook['Results']

    # Define light red fill for abnormal rows
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

    # Find the IS_OUTLIER column index
    header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    is_outlier_col_idx = header_row.index('IS_OUTLIER') + 1 if 'IS_OUTLIER' in header_row else None

    if is_outlier_col_idx:
        # Iterate through rows and apply formatting
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            # Check if IS_OUTLIER is ABNORMAL
            if row[is_outlier_col_idx - 1].value == 'ABNORMAL':
                # Apply red fill to all cells in the row
                for cell in row:
                    cell.fill = red_fill

    # Save to buffer
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output
